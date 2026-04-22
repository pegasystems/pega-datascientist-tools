"""
Impact Analyzer — Transparent Replica
======================================
Mirrors Pega's Impact Analyzer UI layout:
  - Overall Health gauge with formula breakdown
  - Active / Inactive tabs
  - 2-column grid of experiment cards
  - Inside each card: live CI chart + value/engagement lift metrics
  - Below each card: full formula transparency
"""
from __future__ import annotations

import json
import math
import io
import time
from typing import Any

import plotly.graph_objects as go
import requests
import streamlit as st

from ia_engine import (
    EXPERIMENTS,
    Z_95,
    ArmMetrics,
    HealthScore,
    TestResult,
    as_num,
    as_percent,
    build_timeseries,
    calculate_results,
    compute_health,
    parse_payload,
    validate_event,
)
from sample_data import SAMPLE_PAYLOAD, GENERATOR_PROFILES, DEFAULT_PROFILE, generate_batch

# -- page config -----------------------------------------------------------
st.set_page_config(
    page_title="Impact Analyzer - Transparent",
    
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- colour tokens ---------------------------------------------------------
PEGA_BLUE    = "#1F6BFF"
PEGA_GREEN   = "#2ca02c"
PEGA_RED     = "#d62728"
PEGA_ORANGE  = "#f39c12"
PEGA_GREY    = "#8795A1"
PEGA_TEXT     = "#1B2A4A"
C_TEST        = "#1f77b4"
C_CTRL        = "#ff7f0e"
C_BAND        = "rgba(31,119,180,0.12)"


# -- Excel export helper ---------------------------------------------------
def build_excel(results: list[TestResult]) -> bytes:
    """Build a multi-sheet Excel workbook from the computed results."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    red_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Experiment", "Status",
        "Test Impressions", "Test Accepts", "Test Accept Rate",
        "Control Impressions", "Control Accepts", "Control Accept Rate",
        "Engagement Lift", "Eng. SE (rounded 4dp)", "Eng. Significant",
        "Action Value (Test)", "Action Value (Control)",
        "VPI Test", "VPI Control",
        "Value Lift", "Value Lift SE", "Value Lift Significant",
        "Note",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        exp_name = r.definition.get("experimentName", r.definition.get("experimentKey", ""))
        status = "Active" if r.significant is not None else "Inactive"
        values = [
            exp_name, status,
            r.test.impressions, r.test.accepts, r.accept_rate_test,
            r.control.impressions, r.control.accepts, r.accept_rate_control,
            r.lift, r.lift_se_rounded4, r.significant,
            r.action_value_test, r.action_value_control,
            r.vpi_test, r.vpi_control,
            r.value_lift, r.value_lift_se, r.value_lift_significant,
            r.note or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if isinstance(val, float) and val is not None:
                cell.number_format = "0.000000"
            if col_idx == 11 and val is True:  # Eng significant
                cell.fill = green_fill
            elif col_idx == 11 and val is False:
                cell.fill = red_fill
            if col_idx == 18 and val is True:  # Value significant
                cell.fill = green_fill
            elif col_idx == 18 and val is False:
                cell.fill = red_fill

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(results) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # ── Sheet 2: Formula Steps ────────────────────────────────────────
    ws2 = wb.create_sheet("Formula Steps")
    formula_headers = ["Experiment", "Step", "Formula / Calculation"]
    for col_idx, h in enumerate(formula_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for r in results:
        exp_name = r.definition.get("experimentName", r.definition.get("experimentKey", ""))
        for step_name, step_value in r.formulas.items():
            ws2.cell(row=row, column=1, value=exp_name).border = thin_border
            ws2.cell(row=row, column=2, value=step_name).border = thin_border
            ws2.cell(row=row, column=3, value=step_value).border = thin_border
            row += 1

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 80

    # ── Sheet 3: Arm Details ──────────────────────────────────────────
    ws3 = wb.create_sheet("Arm Details")
    arm_headers = [
        "Experiment", "Arm",
        "Impressions", "Accepts", "Accepted Value",
        "Accept Rate", "SE", "Action Value", "VPI", "VPI Variance", "VPI SE",
    ]
    for col_idx, h in enumerate(arm_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for r in results:
        exp_name = r.definition.get("experimentName", r.definition.get("experimentKey", ""))
        for arm_label, arm, rate, se, av, vpi, vpi_var, vpi_se in [
            ("Test", r.test, r.accept_rate_test, r.se_test,
             r.action_value_test, r.vpi_test, r.vpi_var_test, r.vpi_se_test),
            ("Control", r.control, r.accept_rate_control, r.se_control,
             r.action_value_control, r.vpi_control, r.vpi_var_control, r.vpi_se_control),
        ]:
            vals = [exp_name, arm_label, arm.impressions, arm.accepts,
                    arm.accepted_value, rate, se, av, vpi, vpi_var, vpi_se]
            for col_idx, val in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = "0.000000"
            row += 1

    for col_idx in range(1, len(arm_headers) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 18

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -- inject Pega-like CSS -------------------------------------------------
st.markdown("""
<style>
.stApp { background-color: #F4F6F9; }
[data-testid="stHeader"] { background: #1B2A4A; }
.pega-card {
    background: #FFFFFF; border: 1px solid #E0E4EA; border-radius: 8px;
    padding: 20px; margin-bottom: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
}
.pega-card h4 { color: #1B2A4A; font-size: 15px; margin-bottom: 4px; }
.pega-badge-active {
    display: inline-block; background: #2ca02c; color: white; font-size: 11px;
    font-weight: 700; padding: 2px 10px; border-radius: 3px; vertical-align: middle;
    margin-left: 8px;
}
.pega-badge-inactive {
    display: inline-block; background: #8795A1; color: white; font-size: 11px;
    font-weight: 700; padding: 2px 10px; border-radius: 3px; vertical-align: middle;
    margin-left: 8px;
}
.pega-header {
    background: #1B2A4A; color: white; padding: 16px 24px; border-radius: 0;
    margin: -1rem -1rem 1.5rem -1rem;
}
.formula-box {
    background: #F0F4FA; border-left: 4px solid #1F6BFF; padding: 12px 16px;
    font-family: 'SF Mono', 'Fira Code', monospace; font-size: 13px;
    line-height: 1.7; border-radius: 4px; margin: 8px 0;
    white-space: pre-wrap; color: #1B2A4A;
}
.metric-label { font-size: 11px; color: #8795A1; text-transform: uppercase; letter-spacing: 0.5px; }
.metric-value { font-size: 22px; font-weight: 700; color: #1B2A4A; }
.metric-value.positive { color: #2ca02c; }
.metric-value.negative { color: #d62728; }
.metric-value.neutral  { color: #8795A1; }
.section-divider { border-top: 1px solid #E0E4EA; margin: 16px 0; }
.exp-header {
    display: flex; justify-content: space-between; align-items: center;
    background: #F8FAFC; border-bottom: 2px solid #1F6BFF;
    padding: 12px 16px; border-radius: 6px 6px 0 0; margin-bottom: 8px;
}
.exp-title { font-size: 16px; font-weight: 700; color: #1B2A4A; }
.exp-metrics-strip { display: flex; gap: 24px; align-items: center; }
.exp-metric { display: flex; flex-direction: column; align-items: center; }
.exp-metric-lbl { font-size: 10px; color: #8795A1; text-transform: uppercase; letter-spacing: 0.5px; }

/* Style the Lift details expander header with banner color */
details[data-testid="stExpander"] summary:has(span:where([data-testid="stMarkdownContainer"] p)) {
}
div[data-testid="stExpander"]:has(p:contains("Lift details")) > details > summary {
    background: #1B2A4A !important;
    color: white !important;
    border-radius: 4px;
    padding: 10px 16px;
}
.lift-expander-wrap > div[data-testid="stExpander"] > details > summary {
    background: #1B2A4A !important;
    color: white !important;
    border-radius: 6px;
    padding: 10px 16px;
    font-weight: 700;
}
.lift-expander-wrap > div[data-testid="stExpander"] > details > summary span,
.lift-expander-wrap > div[data-testid="stExpander"] > details > summary p {
    color: white !important;
}
.lift-expander-wrap > div[data-testid="stExpander"] > details > summary svg {
    fill: white !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)


# -- session state ---------------------------------------------------------
def _init_state() -> None:
    defaults: dict[str, Any] = {
        "payload_text": SAMPLE_PAYLOAD,
        "events": parse_payload(json.loads(SAMPLE_PAYLOAD)),
        "cursor": 0,
        "live_ingested": 0,
        "auto_refresh": False,
        "refresh_interval": 3,
        "timeframe": "Last 7 days",
        "channel": "All channels",
        # Generator state
        "gen_running": False,
        "gen_counter": 0,
        "gen_batches_sent": 0,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# -- bridge helpers --------------------------------------------------------
def _poll_bridge(base_url: str, cursor: int, limit: int = 1000):
    url = f"{base_url.rstrip('/')}/api/events"
    r = requests.get(url, params={"cursor": cursor, "limit": limit}, timeout=10)
    r.raise_for_status()
    data = r.json()
    records = data.get("events", [])
    next_cur = int(data.get("nextCursor", cursor))
    total = int(data.get("total", 0))
    events: list[dict] = []
    for rec in records:
        payload = rec.get("payload") if isinstance(rec, dict) else None
        if isinstance(payload, dict):
            try:
                events.append(validate_event(payload))
            except ValueError:
                continue
    return events, next_cur, total


# -- chart builders --------------------------------------------------------

def _short(title: str) -> str:
    mapping = {
        "How is Next-Best-Action performing against a random relevant action?": "NBA vs Random relevant action",
        "How is Next-Best-Action performing against arbitrating by propensity-only?": "NBA vs Propensity-only",
        "How is Next-Best-Action performing against arbitrating with no levers?": "NBA vs No levers",
        "How is adaptive model selection performing against random propensity?": "Adaptive model vs Random propensity",
        "How is Next-Best-Action performing against applying only eligibility criteria?": "NBA vs Eligibility only",
    }
    return mapping.get(title, title[:50])


def chart_health_gauge(health: HealthScore) -> go.Figure:
    val = health.score * 100
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=val,
        number=dict(suffix="%", font=dict(size=28, color=PEGA_TEXT)),
        gauge=dict(
            axis=dict(range=[0, 100], tickwidth=1, tickcolor=PEGA_GREY,
                      tickfont=dict(color=PEGA_GREY, size=10)),
            bar=dict(color=health.color, thickness=0.3),
            bgcolor="#E8ECF1", borderwidth=0,
            steps=[
                dict(range=[0, 35], color="#FADBD8"),
                dict(range=[35, 70], color="#FDEBD0"),
                dict(range=[70, 100], color="#D5F5E3"),
            ],
            threshold=dict(line=dict(color=PEGA_TEXT, width=3), thickness=0.8, value=val),
        ),
    ))
    fig.update_layout(
        height=220, margin=dict(l=30, r=30, t=30, b=10),
        paper_bgcolor="rgba(0,0,0,0)", font=dict(color=PEGA_TEXT),
    )
    return fig


def chart_lift_with_ci(result: TestResult, timeseries_snaps: list, lift_type: str = "engagement") -> go.Figure:
    fig = go.Figure()
    if not timeseries_snaps:
        fig.add_annotation(
            text="Awaiting data...", xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False, font=dict(color=PEGA_GREY, size=13),
        )
        fig.update_layout(height=200, paper_bgcolor="rgba(0,0,0,0)",
                          plot_bgcolor="#F8F9FB", font=dict(color=PEGA_TEXT))
        return fig

    xs = [s.event_index for s in timeseries_snaps]

    if lift_type == "value":
        lifts = [s.value_lift * 100 if s.value_lift is not None else None for s in timeseries_snaps]
        ci_lo = [(s.value_ci_lower * 100) if s.value_ci_lower is not None else None for s in timeseries_snaps]
        ci_hi = [(s.value_ci_upper * 100) if s.value_ci_upper is not None else None for s in timeseries_snaps]
        label = "Value Lift %"
    else:
        lifts = [s.lift * 100 if s.lift is not None else None for s in timeseries_snaps]
        ci_lo = [(s.ci_lower * 100) if s.ci_lower is not None else None for s in timeseries_snaps]
        ci_hi = [(s.ci_upper * 100) if s.ci_upper is not None else None for s in timeseries_snaps]
        label = "Lift %"

    valid = [(x, lo, hi) for x, lo, hi in zip(xs, ci_lo, ci_hi) if lo is not None]
    if valid:
        bx = [v[0] for v in valid]
        blo = [v[1] for v in valid]
        bhi = [v[2] for v in valid]
        fig.add_trace(go.Scatter(
            x=bx + bx[::-1], y=bhi + blo[::-1],
            fill="toself", fillcolor=C_BAND, line=dict(width=0),
            hoverinfo="skip", name="95% CI",
        ))

    fig.add_trace(go.Scatter(
        x=xs, y=lifts, mode="lines",
        line=dict(color=PEGA_BLUE, width=2.5), name=label,
        hovertemplate="Event %{x}<br>Lift: %{y:.4f}%<extra></extra>",
    ))
    fig.add_hline(y=0, line_width=1, line_dash="dash", line_color=PEGA_GREY, opacity=0.5)
    fig.update_layout(
        height=200, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="#FAFBFD",
        font=dict(color=PEGA_TEXT, size=11),
        margin=dict(l=40, r=10, t=10, b=30),
        xaxis=dict(title="Events", gridcolor="#EEF0F4"),
        yaxis=dict(title=label, gridcolor="#EEF0F4", zeroline=False),
        legend=dict(orientation="h", y=1.12, font=dict(size=10)),
        showlegend=True,
    )
    return fig


def chart_ci_bar(result: TestResult, lift_type: str = "engagement") -> go.Figure:
    fig = go.Figure()
    if lift_type == "value":
        lift_val = result.value_lift
        se_val = result.value_lift_se
    else:
        lift_val = result.lift
        se_val = result.lift_se

    if lift_val is None or se_val is None:
        fig.add_annotation(text="---", xref="paper", yref="paper", x=0.5, y=0.5,
                           showarrow=False, font=dict(color=PEGA_GREY, size=18))
    else:
        lift_pct = lift_val * 100
        se_pct = se_val * 100
        lo = lift_pct - Z_95 * se_pct
        hi = lift_pct + Z_95 * se_pct
        color = PEGA_GREEN if lift_pct >= 0 else PEGA_RED

        fig.add_shape(type="rect", x0=lo, x1=hi, y0=-0.3, y1=0.3,
                      fillcolor=C_BAND, line_width=0)
        fig.add_trace(go.Scatter(
            x=[lo, hi], y=[0, 0], mode="lines+markers",
            marker=dict(symbol="line-ns-open", size=14, color=PEGA_BLUE, line_width=2),
            line=dict(color=PEGA_BLUE, width=2), showlegend=False,
        ))
        fig.add_trace(go.Scatter(
            x=[lift_pct], y=[0], mode="markers",
            marker=dict(symbol="diamond", size=12, color=color,
                        line=dict(color="white", width=1)),
            showlegend=False,
            hovertemplate=f"Lift: {lift_pct:+.4f}%<br>CI: [{lo:.4f}%, {hi:.4f}%]<extra></extra>",
        ))
        fig.add_vline(x=0, line_width=1, line_dash="dash", line_color=PEGA_GREY, opacity=0.5)

    fig.update_layout(
        height=80, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="#FAFBFD",
        font=dict(color=PEGA_TEXT, size=10),
        margin=dict(l=40, r=10, t=5, b=20),
        yaxis=dict(showticklabels=False, zeroline=False, showgrid=False),
        xaxis=dict(title="Lift %", gridcolor="#EEF0F4"),
    )
    return fig


# -- formula rendering helpers ---------------------------------------------

def _render_formula_box(lines):
    """Render plain-text computation steps in a styled box."""
    if isinstance(lines, str):
        lines = [lines]
    content = "\n".join(lines)
    st.markdown(f'<div class="formula-box">{content}</div>', unsafe_allow_html=True)


def _render_formula(latex: str, description: str = "", substitution: str = ""):
    """Render a formula as: optional description, LaTeX equation, optional substituted values."""
    if description:
        st.markdown(f"<p style='color:#5A6B80;font-size:13px;margin-bottom:2px'>{description}</p>",
                    unsafe_allow_html=True)
    st.latex(latex)
    if substitution:
        st.markdown(
            f'<div class="formula-box" style="margin-top:-8px">{substitution}</div>',
            unsafe_allow_html=True,
        )


def _lift_class(lift):
    if lift is None:
        return "neutral"
    return "positive" if lift >= 0 else "negative"


# -- sidebar ---------------------------------------------------------------
def _render_sidebar():
    with st.sidebar:
        st.markdown("## Data Source")
        source = st.radio(
            "Mode",
            ["Sample data", "Live generator", "File upload (PDC/VBD)", "Live bridge"],
            horizontal=False,
        )

        bridge_url = ""
        limit = 1000

        # ── Sample data ──────────────────────────────────────────────
        if source == "Sample data":
            st.caption("Pre-loaded 30 events across all 5 experiments.")
            if st.button("Reload sample", use_container_width=True):
                st.session_state.payload_text = SAMPLE_PAYLOAD
                st.session_state.events = parse_payload(json.loads(SAMPLE_PAYLOAD))
                st.session_state.gen_running = False
                st.toast("Sample reloaded")

        # ── Live generator ───────────────────────────────────────────
        elif source == "Live generator":
            st.markdown("### Simulate live data")
            profile = st.selectbox(
                "Scenario",
                list(GENERATOR_PROFILES.keys()),
                index=0,
            )
            st.caption(GENERATOR_PROFILES[profile]["description"])

            batch_size = st.slider(
                "Impressions per experiment per tick",
                min_value=10, max_value=500, value=100, step=10,
            )
            gen_interval = st.slider(
                "Tick interval (seconds)",
                min_value=1, max_value=15, value=2,
            )

            g1, g2, g3 = st.columns(3)
            if g1.button("Start", use_container_width=True):
                st.session_state.gen_running = True
                st.session_state.events = []  # fresh start
                st.session_state.gen_counter = 0
                st.session_state.gen_batches_sent = 0
                st.session_state.auto_refresh = True
                st.session_state.refresh_interval = gen_interval
                st.session_state._gen_profile = profile
                st.session_state._gen_batch_size = batch_size
                st.toast("Generator started — watch the metrics update!", )
            if g2.button("Pause", use_container_width=True):
                st.session_state.gen_running = False
                st.session_state.auto_refresh = False
                st.toast("Generator paused", )
            if g3.button("Reset", use_container_width=True):
                st.session_state.gen_running = False
                st.session_state.auto_refresh = False
                st.session_state.events = []
                st.session_state.gen_counter = 0
                st.session_state.gen_batches_sent = 0
                st.toast("Generator reset", )

            st.markdown("---")
            st.caption(f"Status: **{'Running' if st.session_state.gen_running else 'Stopped'}**")
            st.caption(f"Batches sent: **{st.session_state.gen_batches_sent}**")
            st.caption(f"Total events: **{len(st.session_state.events):,}**")

            # Generate one batch if running
            if st.session_state.gen_running:
                new_events = generate_batch(
                    batch_size=st.session_state.get("_gen_batch_size", batch_size),
                    profile_name=st.session_state.get("_gen_profile", profile),
                    counter=st.session_state.gen_counter,
                )
                parsed = parse_payload(new_events)
                st.session_state.events.extend(parsed)
                st.session_state.gen_counter += batch_size * 5  # 5 experiments
                st.session_state.gen_batches_sent += 1

        # ── File upload (PDC JSON / VBD) ─────────────────────────────
        elif source == "File upload (PDC/VBD)":
            st.markdown("### Upload Pega export")
            st.caption(
                "Upload a **PDC JSON** export or a **VBD / Scenario Planner Actuals** "
                "file (CSV, Parquet, JSON). The app will convert it to events."
            )
            uploaded = st.file_uploader(
                "Choose file",
                type=["json", "csv", "parquet", "ndjson", "zip"],
            )
            if uploaded is not None and st.button("Load file", use_container_width=True):
                try:
                    events = _load_uploaded_file(uploaded)
                    st.session_state.events = events
                    st.session_state.gen_running = False
                    st.toast(f"Loaded {len(events):,} events from {uploaded.name}", )
                except Exception as e:
                    st.error(f"Error loading file: {e}")

        # ── Live bridge ──────────────────────────────────────────────
        elif source == "Live bridge":
            bridge_url = st.text_input("Bridge URL", value="http://127.0.0.1:8787")
            limit = st.number_input("Events per poll", 1, 5000, 1000, 100)
            col1, col2 = st.columns(2)
            if col1.button("Fetch", use_container_width=True):
                try:
                    new_ev, nc, total = _poll_bridge(bridge_url, st.session_state.cursor, int(limit))
                    st.session_state.cursor = nc
                    if new_ev:
                        st.session_state.events.extend(new_ev)
                        st.session_state.live_ingested += len(new_ev)
                    st.toast(f"Fetched {len(new_ev)} events (total: {total})")
                except Exception as exc:
                    st.error(f"Bridge error: {exc}")
            if col2.button("Reset", use_container_width=True):
                st.session_state.cursor = 0
                st.session_state.live_ingested = 0

        st.markdown("---")
        st.markdown("### Auto-refresh")
        auto = st.toggle("Enable", value=st.session_state.auto_refresh)
        st.session_state.auto_refresh = auto
        interval = st.slider("Interval (s)", 1, 30, st.session_state.refresh_interval)
        st.session_state.refresh_interval = interval

        st.markdown("---")
        st.caption(f"Cursor: **{st.session_state.cursor}**")
        st.caption(f"Live ingested: **{st.session_state.live_ingested}**")
        st.caption(f"Total events: **{len(st.session_state.events):,}**")

    return bridge_url, int(limit), auto, interval


def _load_uploaded_file(uploaded) -> list[dict[str, Any]]:
    """Convert an uploaded PDC JSON, VBD NDJSON, or event array to event list."""
    name = uploaded.name.lower()

    if name.endswith(".json") or name.endswith(".ndjson"):
        raw_bytes = uploaded.read()
        text = raw_bytes.decode("utf-8", errors="replace").strip()

        # Detect NDJSON (line-delimited JSON) — first char is '{' and multiple lines
        if text.startswith("{") and "\n" in text:
            rows = []
            for line in text.splitlines():
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
            # Check if it looks like VBD/SPActuals (has pyOutcome field)
            if rows and "pyOutcome" in rows[0]:
                return _convert_vbd_ndjson(rows)
            # Otherwise treat each line as a raw event
            return parse_payload(rows)

        # Standard JSON (single object or array)
        raw = json.loads(text)
        # Try PDC format first
        if isinstance(raw, dict) and "pxResults" in raw:
            return _convert_pdc_json(raw)
        # Else treat as raw event array
        if isinstance(raw, list):
            return parse_payload(raw)
        raise ValueError("Unrecognized JSON format — expected PDC export, VBD NDJSON, or event array.")

    # For CSV/Parquet/etc we'd need polars — provide a helpful message
    raise ValueError(
        f"File type '{name.split('.')[-1]}' requires pdstools. "
        f"Use ImpactAnalyzer.from_vbd() for VBD files, or export as JSON."
    )


def _convert_vbd_ndjson(rows: list[dict]) -> list[dict[str, Any]]:
    """Convert VBD / Scenario Planner Actuals NDJSON rows into pdc_summary events.

    Aggregates per DATE per experiment so timeseries charts have multiple points.

    VBD format (one row per action × outcome × date):
    - Rows WITHOUT MktType/pyReason  → total counts (test + control combined)
    - Rows WITH pyReason="Control"   → control-only counts
    - Test counts = total - control
    """
    from collections import defaultdict

    experiments_map = {
        "NBAHealth_NBAPrioritization": "nba_random_relevant_action",
        "NBAHealth_PropensityPriority": "nba_propensity_only",
        "NBAHealth_LeverPriority": "nba_no_levers",
        "NBAHealth_ModelControl_2": "adaptive_vs_random_propensity",
        "NBAHealth_ModelControl": "adaptive_vs_random_propensity",
        "NBAHealth_EngagementPolicy": "nba_eligibility_only",
    }
    control_group_values = {"GA", "CG_IW", "CG_RS", "UG"}

    def _extract_date(row: dict) -> str:
        ts = row.get("pxOutcomeTime", "")
        return ts[:8] if len(ts) >= 8 else "00000000"

    def _empty_bucket():
        return {"impr_ctrl": 0, "acc_ctrl": 0, "val_ctrl": 0.0}

    # Per-date aggregation: {date: {"_total": {impr, acc, val}, exp_key: {impr_ctrl, acc_ctrl, val_ctrl}}}
    daily_total: dict[str, dict[str, int | float]] = defaultdict(lambda: {"impr": 0, "acc": 0, "val": 0.0})
    daily_ctrl: dict[str, dict[str, dict[str, int | float]]] = defaultdict(lambda: defaultdict(_empty_bucket))

    for row in rows:
        outcome = row.get("pyOutcome", "")
        if outcome not in ("Impression", "Accepted"):
            continue

        count = int(row.get("AggregateCount", 1))
        value = float(row.get("Value", 0))
        py_reason = row.get("pyReason", "")
        mkt_value = row.get("MktValue", "")
        date = _extract_date(row)

        if py_reason == "Control":
            if mkt_value in experiments_map:
                exp_key = experiments_map[mkt_value]
            elif mkt_value in control_group_values:
                exp_key = "nba_random_relevant_action"
            else:
                continue
            bucket = daily_ctrl[date][exp_key]
            if outcome == "Impression":
                bucket["impr_ctrl"] += count
            else:
                bucket["acc_ctrl"] += count
                bucket["val_ctrl"] += count * value
        else:
            tot = daily_total[date]
            if outcome == "Impression":
                tot["impr"] += count
            else:
                tot["acc"] += count
                tot["val"] += count * value

    # Collect all experiment keys that appear in any date
    all_exp_keys = set()
    for date_ctrl in daily_ctrl.values():
        all_exp_keys.update(date_ctrl.keys())

    if not all_exp_keys:
        raise ValueError(
            "No experiment data found in VBD file. "
            "Expected rows with MktType/MktValue fields (e.g., NBAHealth_NBAPrioritization)."
        )

    # Emit one pdc_summary event per date per experiment, sorted by date
    events: list[dict[str, Any]] = []
    for date in sorted(daily_total.keys()):
        tot = daily_total[date]
        for exp_key in sorted(all_exp_keys):
            ctrl = daily_ctrl[date].get(exp_key, _empty_bucket())
            impr_test = tot["impr"] - ctrl["impr_ctrl"]
            acc_test = tot["acc"] - ctrl["acc_ctrl"]
            val_test = tot["val"] - ctrl["val_ctrl"]

            if impr_test <= 0 and ctrl["impr_ctrl"] <= 0:
                continue

            av_test = val_test / acc_test if acc_test > 0 else 0
            av_ctrl = ctrl["val_ctrl"] / ctrl["acc_ctrl"] if ctrl["acc_ctrl"] > 0 else 0

            events.append({
                "experimentKey": exp_key,
                "channel": "All channels",
                "impressions_test": max(impr_test, 0),
                "accepts_test": max(acc_test, 0),
                "impressions_control": max(int(ctrl["impr_ctrl"]), 0),
                "accepts_control": max(int(ctrl["acc_ctrl"]), 0),
                "accepted_value_test": max(acc_test, 0) * av_test,
                "accepted_value_control": max(int(ctrl["acc_ctrl"]), 0) * av_ctrl,
                "eventType": "pdc_summary",
                "arm": "test",
            })

    if not events:
        raise ValueError(
            "No experiment data found in VBD file. "
            "Expected rows with MktType/MktValue fields (e.g., NBAHealth_NBAPrioritization)."
        )

    return parse_payload(events)


def _convert_pdc_json(data: dict) -> list[dict[str, Any]]:
    """Convert a PDC JSON export into our event format."""
    from datetime import datetime

    experiments_map = {
        "NBAHealth_NBAPrioritization": "nba_random_relevant_action",
        "NBAHealth_PropensityPriority": "nba_propensity_only",
        "NBAHealth_LeverPriority": "nba_no_levers",
        "NBAHealth_ModelControl": "adaptive_vs_random_propensity",
        "NBAHealth_EngagementPolicy": "nba_eligibility_only",
    }

    events: list[dict[str, Any]] = []
    px = data.get("pxResults", [{}])[0].get("pxResults", [])

    for row in px:
        exp_name = row.get("ExperimentName", "")
        exp_key = experiments_map.get(exp_name)
        if not exp_key:
            continue

        channel = row.get("ChannelName", "All channels")
        # Only use "All channels" aggregates to avoid double-counting
        if channel != "All channels":
            continue

        impr_test = int(row.get("Impressions_NBA", 0))
        acc_test = int(row.get("Accepts_NBA", 0))
        impr_ctrl = int(row.get("Impressions_Control", 0))
        acc_ctrl = int(row.get("Accepts_Control", 0))
        eng_lift = float(row.get("EngagementLift", 0))
        val_lift = float(row.get("ValueLift", 0))
        eng_interval = float(row.get("EngagementLiftInterval", 0))
        val_interval = float(row.get("ValueLiftInterval", 0))
        is_sig = bool(row.get("IsSignificant", False))
        color = row.get("ExperimentColor", "Gray")
        heading = row.get("Heading", "")
        guidance = row.get("Guidance", "")

        # Derive different action values for test vs control so value lift differs
        # from engagement lift. Use PDC's ValueLift to back-calculate control AV.
        av_test = 75.0
        if val_lift != 0 and eng_lift != 0:
            # ValueLift = (VPI_t - VPI_c)/VPI_c where VPI = p * AV
            # If AV_t = AV_c then ValueLift = EngLift. To make them differ,
            # solve: val_lift = (p_t*av_t - p_c*av_c)/(p_c*av_c)
            # => av_c = p_t * av_t / (p_c * (1 + val_lift))
            p_t = acc_test / impr_test if impr_test > 0 else 0
            p_c = acc_ctrl / impr_ctrl if impr_ctrl > 0 else 0
            if p_c > 0:
                av_ctrl = (p_t * av_test) / (p_c * (1 + val_lift))
            else:
                av_ctrl = av_test
        else:
            av_ctrl = av_test

        # Create a single summary event per row (not one per impression!)
        events.append({
            "experimentKey": exp_key,
            "channel": channel,
            "impressions_test": impr_test,
            "accepts_test": acc_test,
            "impressions_control": impr_ctrl,
            "accepts_control": acc_ctrl,
            "accepted_value_test": acc_test * av_test,
            "accepted_value_control": acc_ctrl * av_ctrl,
            "engagement_lift": eng_lift,
            "value_lift": val_lift,
            "engagement_lift_interval": eng_interval,
            "value_lift_interval": val_interval,
            "is_significant": is_sig,
            "color": color,
            "heading": heading,
            "guidance": guidance,
            "eventType": "pdc_summary",
            "arm": "test",
        })

    return parse_payload(events)


# -- experiment card -------------------------------------------------------
def _render_experiment_card(result, timeseries_snaps, card_index):
    title = result.definition["title"]
    short = _short(title)
    is_active = result.test.impressions > 0 or result.control.impressions > 0

    badge = ('<span class="pega-badge-active">ACTIVE</span>' if is_active
             else '<span class="pega-badge-inactive">INACTIVE</span>')

    st.markdown(
        f'<div class="pega-card"><h4>{short} {badge}</h4></div>',
        unsafe_allow_html=True,
    )

    # Significance callout
    if result.significant is True:
        st.success("Statistically significant at 95% confidence")
    elif result.significant is False:
        st.warning("Not yet significant — more data needed")
    else:
        st.info("Not enough data for insights. Continue monitoring.")

    # Metrics row
    m1, m2, m3, m4 = st.columns(4)
    vl_cls = _lift_class(result.lift)
    vl_val_cls = _lift_class(result.value_lift)
    m1.markdown(f'<div class="metric-label">ENG. LIFT</div>'
                f'<div class="metric-value {vl_cls}">{as_percent(result.lift)}</div>',
                unsafe_allow_html=True)
    m2.markdown(f'<div class="metric-label">VALUE LIFT</div>'
                f'<div class="metric-value {vl_val_cls}">{as_percent(result.value_lift)}</div>',
                unsafe_allow_html=True)
    m3.markdown(f'<div class="metric-label">TEST n</div>'
                f'<div class="metric-value">{result.test.impressions:,}</div>',
                unsafe_allow_html=True)
    m4.markdown(f'<div class="metric-label">CTRL n</div>'
                f'<div class="metric-value">{result.control.impressions:,}</div>',
                unsafe_allow_html=True)

    # Arm comparison table
    arm_data = {
        "": ["Impressions", "Accepts", "Accept Rate"],
        f"Test — {result.definition['testLabel']}": [
            f"{result.test.impressions:,}",
            f"{result.test.accepts:,}",
            as_percent(result.accept_rate_test),
        ],
        f"Control — {result.definition['controlLabel']}": [
            f"{result.control.impressions:,}",
            f"{result.control.accepts:,}",
            as_percent(result.accept_rate_control),
        ],
    }
    st.table(arm_data)

    # Charts — tabbed: Engagement Lift / Value Lift
    tab_eng, tab_val = st.tabs(["Engagement Lift", "Value Lift"])

    with tab_eng:
        st.plotly_chart(
            chart_lift_with_ci(result, timeseries_snaps, lift_type="engagement"),
            use_container_width=True, key=f"lift_ci_eng_{card_index}",
        )
        st.caption("95 % Confidence Interval — Engagement Lift")
        st.plotly_chart(
            chart_ci_bar(result, lift_type="engagement"),
            use_container_width=True, key=f"ci_bar_eng_{card_index}",
        )

    with tab_val:
        st.plotly_chart(
            chart_lift_with_ci(result, timeseries_snaps, lift_type="value"),
            use_container_width=True, key=f"lift_ci_val_{card_index}",
        )
        st.caption("95 % Confidence Interval — Value Lift")
        st.plotly_chart(
            chart_ci_bar(result, lift_type="value"),
            use_container_width=True, key=f"ci_bar_val_{card_index}",
        )

    # Lift details (formulas)
    st.markdown('<div class="lift-expander-wrap">', unsafe_allow_html=True)
    with st.expander("**Lift details — show formula steps**"):
        st.markdown("##### Step 1 — Accept Rates")
        _render_formula(
            r"p = \frac{\text{Accepts}}{\text{Impressions}}",
            "Accept rate for each arm",
            f"p_test = {result.test.accepts} / {result.test.impressions} = {as_num(result.accept_rate_test, 10)}\n"
            f"p_ctrl = {result.control.accepts} / {result.control.impressions} = {as_num(result.accept_rate_control, 10)}",
        )

        st.markdown("##### Step 2 — Standard Errors")
        _render_formula(
            r"\text{SE}(p) = \sqrt{\frac{p\,(1-p)}{n}}",
            "Binomial standard error",
            f"SE_test = {as_num(result.se_test, 10)}\n"
            f"SE_ctrl = {as_num(result.se_control, 10)}",
        )

        st.markdown("##### Step 3 — Engagement Lift")
        _render_formula(
            r"\text{Lift} = \frac{p_{\text{test}} - p_{\text{ctrl}}}{p_{\text{ctrl}}}",
            "",
            f"Lift = ({as_num(result.accept_rate_test,10)} − {as_num(result.accept_rate_control,10)}) / {as_num(result.accept_rate_control,10)} = {as_num(result.lift, 10)}"
            if result.lift is not None else "Lift = N/A (control rate is 0)",
        )

        st.markdown("##### Step 4 — Confidence Interval (Delta Method)")
        _render_formula(
            r"\text{SE}_{\text{lift}} = \frac{1}{p_c}\,\sqrt{\text{SE}_t^2 + \left(\frac{p_t}{p_c}\right)^{\!2} \text{SE}_c^2}",
            "",
            f"SE(Lift) = {as_num(result.lift_se, 10)}" if result.lift_se is not None else "SE(Lift) = N/A",
        )
        _render_formula(
            r"\text{CI}_{95\%} = \text{Lift} \pm 1.96 \times \text{SE}_{\text{lift}}",
        )
        if result.lift is not None and result.lift_se is not None:
            lo = result.lift - Z_95 * result.lift_se
            hi = result.lift + Z_95 * result.lift_se
            sig_text = "YES — CI does not contain 0" if result.significant else "NO — CI contains 0"
            _render_formula_box([
                f"= [{lo:.10f},  {hi:.10f}]",
                f"Statistically significant?  {sig_text}",
            ])

        st.markdown("##### Step 5 — Value Lift (VPI-based)")
        _render_formula(
            r"\text{VPI} = p \times \text{ActionValue}",
            "Value Per Impression",
            f"AV_test = {as_num(result.action_value_test, 4)}  |  AV_ctrl = {as_num(result.action_value_control, 4)}\n"
            f"VPI_test = {as_num(result.accept_rate_test, 10)} x {as_num(result.action_value_test, 4)} = {as_num(result.vpi_test, 10)}\n"
            f"VPI_ctrl = {as_num(result.accept_rate_control, 10)} x {as_num(result.action_value_control, 4)} = {as_num(result.vpi_control, 10)}",
        )
        _render_formula(
            r"\text{Var}(\text{VPI}) = p\,(1-p) \times \text{AV}^2",
            "Bernoulli per-observation variance",
            f"Var_test = {as_num(result.vpi_var_test, 10)}\n"
            f"Var_ctrl = {as_num(result.vpi_var_control, 10)}",
        )
        _render_formula(
            r"\text{SE}(\text{VPI}) = \sqrt{\frac{\text{Var}}{n}}",
            "",
            f"SE_vpi_test = {as_num(result.vpi_se_test, 10)}\n"
            f"SE_vpi_ctrl = {as_num(result.vpi_se_control, 10)}",
        )
        _render_formula(
            r"\text{ValueLift} = \frac{\text{VPI}_{\text{test}} - \text{VPI}_{\text{ctrl}}}{\text{VPI}_{\text{ctrl}}}",
            "",
            f"ValueLift = {as_num(result.value_lift, 10)}" if result.value_lift is not None
            else "ValueLift = N/A (control VPI is 0)",
        )
        if result.value_lift is not None and result.value_lift_se is not None:
            vl_lo = result.value_lift - Z_95 * result.value_lift_se
            vl_hi = result.value_lift + Z_95 * result.value_lift_se
            vl_sig_text = "YES — CI does not contain 0" if result.value_lift_significant else "NO — CI contains 0"
            _render_formula_box([
                f"SE(ValueLift) = {as_num(result.value_lift_se, 10)}",
                f"95% CI = [{vl_lo:.10f},  {vl_hi:.10f}]",
                f"Statistically significant?  {vl_sig_text}",
            ])
    st.markdown('</div>', unsafe_allow_html=True)


# -- main ------------------------------------------------------------------
def main() -> None:
    _init_state()
    bridge_url, limit, auto_refresh, interval = _render_sidebar()

    # Auto-refresh (bridge or generator)
    if auto_refresh:
        if bridge_url:
            try:
                new_ev, nc, _ = _poll_bridge(bridge_url, st.session_state.cursor, limit)
                if new_ev:
                    st.session_state.cursor = nc
                    st.session_state.events.extend(new_ev)
                    st.session_state.live_ingested += len(new_ev)
            except Exception:
                pass
        time.sleep(interval)
        st.rerun()

    # Header (Pega-style)
    st.markdown(
        '<div class="pega-header">'
        '<span style="font-size:12px;opacity:0.7">PEGA &nbsp; Customer Decision Hub</span>'
        '<h2 style="margin:4px 0 0 0;color:white">Impact Analyzer</h2>'
        '</div>',
        unsafe_allow_html=True,
    )

    # Toolbar: Timeframe / Channels / Apply
    tf_col, ch_col, btn_col, export_col = st.columns([2, 2, 1, 2])
    with tf_col:
        st.session_state.timeframe = st.selectbox(
            "Timeframe", ["Last 7 days", "Last 14 days", "Last 30 days", "All time"],
        )
    with ch_col:
        st.session_state.channel = st.selectbox(
            "Channels", ["All channels", "Web", "Mobile", "Email", "SMS", "Push"],
        )
    with btn_col:
        st.markdown("<br>", unsafe_allow_html=True)
        st.button("Apply", use_container_width=True)
    with export_col:
        st.markdown("<br>", unsafe_allow_html=True)
        # Real export — computed after results are available; use session state
        if st.session_state.get("_excel_bytes"):
            st.download_button(
                "Export Excel file", data=st.session_state["_excel_bytes"],
                file_name="ImpactAnalyzerExport.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.download_button(
                "Export Excel file", data=b"",
                file_name="ImpactAnalyzerExport.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                disabled=True,
            )

    st.caption(f"Available dates: {st.session_state.timeframe} | Channel: {st.session_state.channel}")

    # Payload editor (collapsed)
    with st.expander("Event payload - click to expand"):
        st.session_state.payload_text = st.text_area(
            "payload", st.session_state.payload_text, height=200,
            label_visibility="collapsed",
        )
        p1, p2, p3 = st.columns(3)
        if p1.button("Apply payload", use_container_width=True):
            try:
                st.session_state.events = parse_payload(json.loads(st.session_state.payload_text))
                st.toast(f"Applied {len(st.session_state.events)} events")
            except Exception as e:
                st.error(str(e))
        if p2.button("Load sample", use_container_width=True):
            st.session_state.payload_text = SAMPLE_PAYLOAD
            st.session_state.events = parse_payload(json.loads(SAMPLE_PAYLOAD))
        if p3.button("Clear", use_container_width=True):
            st.session_state.events = []

    # Compute
    events = st.session_state.events
    results = calculate_results(events)
    timeseries = build_timeseries(events)
    health = compute_health(results)

    # Build Excel export data for the download button
    if results:
        st.session_state["_excel_bytes"] = build_excel(results)
    else:
        st.session_state["_excel_bytes"] = None

    # =====================================================================
    # SECTION 1: Overall Health
    # =====================================================================
    st.markdown("### Overall health")

    h_left, h_right = st.columns([1, 2])
    with h_left:
        st.plotly_chart(chart_health_gauge(health), use_container_width=True, key="health_gauge")
    with h_right:
        st.markdown(f"**{health.label}**")
        if health.label == "Not enough data":
            st.markdown(
                "**Not enough data for the selected timeframe.**  \n"
                "Please continue to monitor as Impact Analyzer processes incoming data."
            )
        elif health.label == "Fair":
            st.markdown(
                "Some experiments are collecting data but have not reached "
                "statistical significance yet."
            )
        else:
            st.markdown("All experiments are performing well with sufficient data.")
        if health.reasons:
            for r in health.reasons[:5]:
                st.caption(f"* {r}")

    with st.expander("How is Overall Health computed?"):
        _render_formula_box(health.formula_steps)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # =====================================================================
    # SECTION 2: Active / Inactive tabs with experiment cards
    # =====================================================================
    active = [r for r in results if r.test.impressions > 0 or r.control.impressions > 0]
    inactive = [r for r in results if r.test.impressions == 0 and r.control.impressions == 0]

    tab_active, tab_inactive = st.tabs([
        f"Active ({len(active)})",
        f"Inactive ({len(inactive)})",
    ])

    with tab_active:
        if not active:
            st.info("No active experiments. Feed data to see results.")
        else:
            cols = st.columns(2)
            for idx, r in enumerate(active):
                ts_key = r.definition["key"]
                with cols[idx % 2]:
                    _render_experiment_card(r, timeseries.get(ts_key, []), idx)

    with tab_inactive:
        if not inactive:
            st.success("All experiments are active!")
        else:
            for i, r in enumerate(inactive):
                st.markdown(
                    f'<div class="pega-card">'
                    f'<h4>{r.definition["title"]} '
                    f'<span class="pega-badge-inactive">INACTIVE</span></h4>'
                    f'<p style="color:#8795A1">No data received yet for this experiment.</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    # =====================================================================
    # SECTION 3: Summary table
    # =====================================================================
    st.markdown("---")
    st.markdown("### Summary Table")
    rows = []
    for r in results:
        rows.append({
            "Experiment": _short(r.definition["title"]),
            "Test Impressions": f"{r.test.impressions:,}",
            "Ctrl Impressions": f"{r.control.impressions:,}",
            "Test Rate": as_percent(r.accept_rate_test),
            "Ctrl Rate": as_percent(r.accept_rate_control),
            "Engagement Lift": as_percent(r.lift),
            "SE(Lift)": as_num(r.lift_se_rounded4, 4),
            "Significant": "Yes" if r.significant else ("No" if r.significant is False else "-"),
        })
    st.dataframe(rows, use_container_width=True, hide_index=True)

    # =====================================================================
    # SECTION 4: Full formula reference
    # =====================================================================
    with st.expander("Full formula reference - all Impact Analyzer calculations"):
        st.markdown("#### Engagement Metrics")
        _render_formula(r"p = \frac{\text{Accepts}}{\text{Impressions}}", "Accept Rate")
        _render_formula(r"\text{SE}(p) = \sqrt{\frac{p\,(1-p)}{n}}", "Standard Error (Binomial)")
        _render_formula(r"\text{CI}(p) = z \times \text{SE}(p) \quad\text{where } z = 1.96 \text{ for 95\%}", "Confidence Interval")

        st.markdown("---")
        st.markdown("#### Lift (Relative)")
        _render_formula(r"\text{Lift} = \frac{p_{\text{test}} - p_{\text{ctrl}}}{p_{\text{ctrl}}}")
        _render_formula(
            r"\text{SE}(\text{Lift}) = \frac{1}{p_c}\,\sqrt{\text{SE}_t^2 + \left(\frac{p_t}{p_c}\right)^{\!2} \text{SE}_c^2}",
            "Delta method for ratio of two proportions",
        )
        _render_formula(r"\text{CI}_{95\%}(\text{Lift}) = \text{Lift} \pm 1.96 \times \text{SE}(\text{Lift})")
        _render_formula_box([
            "Significant iff CI does not contain 0:",
            "  (Lift - 1.96*SE > 0)  OR  (Lift + 1.96*SE < 0)",
        ])

        st.markdown("---")
        st.markdown("#### Value Per Impression (VPI)")
        _render_formula(r"\text{VPI} = \frac{\text{Accepts}}{\text{Impressions}} \times \text{AV} = p \times \text{AV}")
        _render_formula(r"\text{Var}(\text{VPI}) = p\,(1-p) \times \text{AV}^2", "Bernoulli per-observation variance")
        _render_formula(r"\text{SE}(\text{VPI}) = \sqrt{\frac{\text{Var}(\text{VPI})}{n}}")
        _render_formula(r"\text{ValueLift} = \frac{\text{VPI}_{\text{test}} - \text{VPI}_{\text{ctrl}}}{\text{VPI}_{\text{ctrl}}}")
        _render_formula(
            r"\text{CI}(\text{ValueLift}) = \frac{z}{\text{VPI}_c}\,\sqrt{\text{SE}_{\text{vpi}_t}^2 + \left(\frac{\text{VPI}_t}{\text{VPI}_c}\right)^{\!2} \text{SE}_{\text{vpi}_c}^2}",
            "Delta-method CI for value lift",
        )

        st.markdown("---")
        st.markdown("#### Required Sample Size")
        _render_formula(r"n = \frac{(z_\alpha + z_\beta)^2 \cdot p\,(1-p)}{\delta^2}")
        _render_formula_box([
            "where  z_alpha = 1.96 (95% confidence)",
            "       z_beta  = 0.84 (80% power)",
            "       delta   = p x MDE   (minimum detectable effect)",
        ])

        st.markdown("---")
        st.markdown("#### Overall Health Score")
        _render_formula(r"\text{Score} = \frac{\text{total\_points}}{3 \times N_{\text{experiments}}}")
        _render_formula_box([
            "Per experiment (max 3 points):",
            "  +1  control arm has impressions",
            "  +1  both arms >= n_min impressions",
            "  +1  statistically significant",
            "",
            "Score >= 70%  ->  Good",
            "Score >= 35%  ->  Fair",
            "Score <  35%  ->  Not enough data",
        ])


if __name__ == "__main__":
    main()
