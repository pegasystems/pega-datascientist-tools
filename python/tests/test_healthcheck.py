import pathlib

import pytest
from openpyxl import load_workbook
from pdstools import ADMDatamart, Prediction, datasets, read_ds_export
from pdstools.utils.report_utils import check_report_for_errors

basePath = pathlib.Path(__file__).parent.parent.parent


def _assert_report_path(actual: pathlib.Path, expected_stem: str) -> None:
    """Assert ``actual`` is the rendered report for ``expected_stem``.

    The output may be either ``<stem>.html`` (when Quarto did not emit a
    resources folder, or when ``full_embed=True``) or ``<stem>.zip`` (when
    the resources folder was bundled). Tests should be agnostic to which.
    """
    actual = pathlib.Path(actual)
    assert actual.parent.resolve() == pathlib.Path(".").resolve()
    assert actual.stem == expected_stem
    assert actual.suffix in {".html", ".zip"}, f"Unexpected extension: {actual.suffix}"
    assert actual.exists()


@pytest.fixture
def sample() -> ADMDatamart:
    return datasets.cdh_sample()


@pytest.fixture
def sample_without_predictor_binning() -> ADMDatamart:
    """Fixture to serve as class to call functions from."""
    # Using from_ds_export automaticaly detects predictor_snapshot.
    model_df = read_ds_export(
        filename="Data-Decision-ADM-ModelSnapshot_pyModelSnapshots_20210526T131808_GMT.zip",
        path=f"{basePath}/data",
    )
    return ADMDatamart(model_df=model_df)


@pytest.fixture
def sample_prediction_data() -> Prediction:
    return Prediction.from_mock_data(days=30)


def test_GenerateHealthCheck(sample: ADMDatamart):
    hc = sample.generate.health_check()
    _assert_report_path(hc, "HealthCheck")
    errors = check_report_for_errors(hc)
    pathlib.Path(hc).unlink()
    assert not pathlib.Path(hc).exists()
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_HealthCheck_NoErrors(sample: ADMDatamart, tmp_path):
    """Test that HealthCheck report generates without errors."""
    hc = sample.generate.health_check(output_dir=tmp_path, name="error_check")
    assert pathlib.Path(hc).exists()

    # Check for errors in the generated HTML
    errors = check_report_for_errors(hc)

    # Clean up
    pathlib.Path(hc).unlink()

    # Assert no errors were found
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


@pytest.mark.slow
def test_HealthCheck_full_embed(sample: ADMDatamart, tmp_path):
    """Test health check file sizes for full_embed options."""
    sizes = {}

    default = sample.generate.health_check(output_dir=tmp_path, name="default")
    sizes["default"] = default.stat().st_size

    print("Generating full_embed...")
    full_embed = sample.generate.health_check(
        output_dir=tmp_path,
        name="full_embed",
        full_embed=True,
    )
    sizes["full_embed"] = full_embed.stat().st_size

    print("Generating cdn...")
    cdn = sample.generate.health_check(
        output_dir=tmp_path,
        name="cdn",
        full_embed=False,
    )
    sizes["cdn"] = cdn.stat().st_size

    size_diff = abs(sizes["default"] - sizes["cdn"]) / sizes["cdn"]
    assert size_diff <= 0.10, f"Default is cdn, file sizes could be slightly different, got {size_diff:.1%} difference"

    full_embed_mb = sizes["full_embed"] / (1024 * 1024)
    assert 50 <= full_embed_mb <= 150, f"Embedded size should be large, got {full_embed_mb:.1f} MB"


def test_ExportTables(sample: ADMDatamart):
    excel, warning_messages = sample.generate.excel_report(predictor_binning=True)
    assert excel == pathlib.Path("./Tables.xlsx")
    assert excel.exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
        "predictors_detail",
        "predictors_overview",
        "predictor_binning",
    ]
    # TODO we could go further and check the size of the sheets
    # spreadsheet = read_excel(excel, sheet_name=None)
    pathlib.Path(excel).unlink()
    assert not pathlib.Path(excel).exists()


def test_ExportTables_NoBinning(sample: ADMDatamart):
    excel, warining_messages = sample.generate.excel_report(predictor_binning=False)
    assert excel == pathlib.Path("./Tables.xlsx")
    assert pathlib.Path(excel).exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
        "predictors_detail",
        "predictors_overview",
    ]
    # TODO we could go further and check the size of the sheets
    # spreadsheet = read_excel(excel, sheet_name=None)
    pathlib.Path(excel).unlink()
    assert not pathlib.Path(excel).exists()


def test_GenerateHealthCheck_ModelDataOnly(
    sample_without_predictor_binning: ADMDatamart,
):
    hc = sample_without_predictor_binning.generate.health_check(name="MyOrg")
    _assert_report_path(hc, "HealthCheck_MyOrg")
    errors = check_report_for_errors(hc)
    pathlib.Path(hc).unlink()
    assert not pathlib.Path(hc).exists()
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_GenerateHealthCheck_PredictionData(
    sample: ADMDatamart,
    sample_prediction_data: Prediction,
):
    hc = sample.generate.health_check(
        prediction=sample_prediction_data,
        name="WithPredictions",
    )
    _assert_report_path(hc, "HealthCheck_WithPredictions")
    errors = check_report_for_errors(hc)
    pathlib.Path(hc).unlink()
    assert not pathlib.Path(hc).exists()
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_ExportTables_ModelDataOnly(sample_without_predictor_binning: ADMDatamart):
    excel, warning_messages = sample_without_predictor_binning.generate.excel_report(
        name="ModelTables.xlsx",
        predictor_binning=True,
    )
    assert excel == pathlib.Path("ModelTables.xlsx")
    assert pathlib.Path(excel).exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
    ]
    # TODO we could go further and check the size of the sheets
    # spreadsheet = read_excel(excel, sheet_name=None)
    pathlib.Path(excel).unlink()


def test_GenerateModelReport(sample: ADMDatamart):
    report = sample.generate.model_reports(
        model_ids=["bd70a915-697a-5d43-ab2c-53b0557c85a0"],
        name="MyOrg",
        only_active_predictors=True,
    )
    _assert_report_path(report, "ModelReport_MyOrg_bd70a915-697a-5d43-ab2c-53b0557c85a0")
    errors = check_report_for_errors(report)
    pathlib.Path(report).unlink()
    assert not pathlib.Path(report).exists()
    assert len(errors) == 0, "Model report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


@pytest.mark.slow
def test_ModelReport_full_embed(sample: ADMDatamart, tmp_path):
    """Test model report file sizes for full_embed options."""
    model_id = "bd70a915-697a-5d43-ab2c-53b0557c85a0"
    sizes = {}

    default = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="default",
    )
    sizes["default"] = default.stat().st_size

    print("Generating full_embed...")
    full_embed = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="full_embed",
        full_embed=True,
    )
    sizes["full_embed"] = full_embed.stat().st_size

    cdn = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="cdn",
        full_embed=False,
    )
    sizes["cdn"] = cdn.stat().st_size

    size_diff = abs(sizes["default"] - sizes["cdn"]) / sizes["cdn"]
    assert size_diff <= 0.10, f"Default is cdn and sizes should be very close, got {size_diff:.1%} difference"

    full_embed_mb = sizes["full_embed"] / (1024 * 1024)
    assert 90 <= full_embed_mb <= 150, f"Embedded size should be large, got {full_embed_mb:.1f} MB"


def test_GenerateHealthCheck_CustomQmdFile(sample: ADMDatamart, tmp_path):
    """Test health_check with custom qmd_file argument"""
    # Create a custom qmd file
    custom_qmd = tmp_path / "custom_health.qmd"
    custom_qmd.write_text("""
---
title: "Custom Health Check Report"
format: html
params:
  title: "Custom Title"
---

# Custom Health Check

This is a custom health check template.

Test parameters: {{< meta params.title >}}
""")

    hc = sample.generate.health_check(name="CustomTemplate", qmd_file=custom_qmd)
    _assert_report_path(hc, "HealthCheck_CustomTemplate")
    pathlib.Path(hc).unlink()
    assert not pathlib.Path(hc).exists()


def test_GenerateModelReport_CustomQmdFile(sample: ADMDatamart, tmp_path):
    """Test model_reports with custom qmd_file argument"""
    # Create a custom qmd file
    custom_qmd = tmp_path / "custom_model.qmd"
    custom_qmd.write_text("""
---
title: "Custom Model Report"
format: html
params:
  title: "Custom Model Title"
  model_id: "test"
---

# Custom Model Report

This is a custom model report template.

Model ID: {{< meta params.model_id >}}
""")

    report = sample.generate.model_reports(
        model_ids=["bd70a915-697a-5d43-ab2c-53b0557c85a0"],
        name="CustomModel",
        qmd_file=custom_qmd,
    )
    _assert_report_path(
        report,
        "ModelReport_CustomModel_bd70a915-697a-5d43-ab2c-53b0557c85a0",
    )
    pathlib.Path(report).unlink()
    assert not pathlib.Path(report).exists()
