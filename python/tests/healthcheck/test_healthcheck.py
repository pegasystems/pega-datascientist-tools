import pathlib
import re
import zipfile
from html.parser import HTMLParser

import pytest
from openpyxl import load_workbook
from pdstools import ADMDatamart, Prediction, datasets, read_ds_export
from pdstools.utils.report_utils import check_report_for_errors

basePath = pathlib.Path(__file__).parent.parent.parent.parent

PLOTLY_CDN_LOAD_RE = re.compile(r"(?:src=|import\s+)[\"']https://cdn\.plot\.ly/")
CSS_URL_RE = re.compile(r"url\(\s*([\"']?)([^\"')]+)\1\s*\)", re.IGNORECASE)
REMOTE_REFERENCE_PREFIXES = ("http://", "https://", "//", "data:", "#", "about:", "javascript:", "mailto:", "tel:")


class ResourceReferenceParser(HTMLParser):
    """Collect live HTML resource references from generated reports."""

    def __init__(self):
        super().__init__()
        self.references: list[tuple[str, str, str]] = []
        self.style_chunks: list[str] = []
        self._in_style = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "style":
            self._in_style = True
        for attr, value in attrs:
            if attr in {"href", "src"} and value:
                self.references.append((tag, attr, value))

    def handle_endtag(self, tag: str) -> None:
        if tag == "style":
            self._in_style = False

    def handle_data(self, data: str) -> None:
        if self._in_style:
            self.style_chunks.append(data)


def _is_local_file_reference(reference: str) -> bool:
    return bool(reference.strip()) and not reference.strip().startswith(REMOTE_REFERENCE_PREFIXES)


def _local_file_references(html: str) -> list[str]:
    parser = ResourceReferenceParser()
    parser.feed(html)
    references = [
        f"{tag}[{attr}]={reference}"
        for tag, attr, reference in parser.references
        if _is_local_file_reference(reference)
    ]
    references.extend(
        f"style[url]={reference.strip()}"
        for style in parser.style_chunks
        for reference in (match.group(2) for match in CSS_URL_RE.finditer(style))
        if _is_local_file_reference(reference)
    )
    return references


def _assert_report_path(actual: pathlib.Path, parent: pathlib.Path, expected_stem: str) -> None:
    """Assert ``actual`` is the rendered report for ``expected_stem`` in ``parent``.

    The output may be either ``<stem>.html`` (when Quarto did not emit a
    resources folder, or when ``full_embed=True``) or ``<stem>.zip`` (when
    the resources folder was bundled). Tests should be agnostic to which.
    """
    actual = pathlib.Path(actual)
    assert actual.parent.resolve() == pathlib.Path(parent).resolve()
    assert actual.stem == expected_stem
    assert actual.suffix in {".html", ".zip"}, f"Unexpected extension: {actual.suffix}"
    assert actual.exists()


def _read_report_html(report_path: pathlib.Path) -> str:
    if report_path.suffix == ".zip":
        with zipfile.ZipFile(report_path) as report_zip:
            html_members = [name for name in report_zip.namelist() if name.endswith(".html")]
            assert len(html_members) == 1
            return report_zip.read(html_members[0]).decode("utf-8")
    return report_path.read_text(encoding="utf-8")


def _assert_plotly_resource_mode(
    cdn_report: pathlib.Path, full_embed_report: pathlib.Path, *, report_label: str
) -> None:
    cdn_html = _read_report_html(cdn_report)
    full_embed_html = _read_report_html(full_embed_report)

    assert PLOTLY_CDN_LOAD_RE.search(cdn_html), f"{report_label} CDN output should load Plotly from the CDN"
    assert not PLOTLY_CDN_LOAD_RE.search(full_embed_html), (
        f"{report_label} full-embed output should not load Plotly from the CDN"
    )
    assert "Plotly.newPlot" in full_embed_html, f"{report_label} full-embed output should still contain Plotly charts"


def _assert_full_embed_size_ratio(sizes: dict[str, int], *, report_label: str, min_ratio: float) -> None:
    ratio = sizes["full_embed"] / sizes["cdn"]
    assert ratio >= min_ratio, (
        f"{report_label} full-embed output should be at least {min_ratio:.1f}x "
        f"the CDN output, got {ratio:.1f}x "
        f"(CDN {sizes['cdn'] / (1024 * 1024):.1f} MB, "
        f"full-embed {sizes['full_embed'] / (1024 * 1024):.1f} MB)"
    )


@pytest.fixture
def sample() -> ADMDatamart:
    return datasets.cdh_sample()


@pytest.fixture
def sample_without_predictor_binning() -> ADMDatamart:
    """Fixture to serve as class to call functions from."""
    # Using from_ds_export automaticaly detects predictor_snapshot.
    model_df = read_ds_export(
        filename="Data-Decision-ADM-ModelSnapshot_pyModelSnapshots_20210526T131808_GMT.zip",
        path=f"{basePath}/data",
    )
    return ADMDatamart(model_df=model_df)


@pytest.fixture
def sample_prediction_data() -> Prediction:
    return Prediction.from_mock_data(days=30)


def test_GenerateHealthCheck(sample: ADMDatamart, tmp_path):
    """Default health-check generation: produces a valid report and contains no errors."""
    hc = sample.generate.health_check(output_dir=tmp_path)
    _assert_report_path(hc, tmp_path, "HealthCheck")
    errors = check_report_for_errors(hc)
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_GenerateHealthCheck_default_cdn_is_djs_copy_safe(sample: ADMDatamart, tmp_path, monkeypatch):
    """Default CDN Health Check can be copied as one standalone HTML file."""
    monkeypatch.chdir(tmp_path)

    hc = sample.generate.health_check(output_type="html")

    assert hc == tmp_path / "HealthCheck.html"
    html = hc.read_text(encoding="utf-8")
    assert PLOTLY_CDN_LOAD_RE.search(html)
    assert "HealthCheck_files/" not in html
    assert not (tmp_path / "HealthCheck.zip").exists()

    upload_dir = tmp_path / "upload"
    upload_dir.mkdir()
    copied_html = upload_dir / "HealthCheck.html"
    copied_html.write_bytes(hc.read_bytes())
    assert _local_file_references(copied_html.read_text(encoding="utf-8")) == []


def test_GenerateHealthCheck_named_cdn_drops_qmd_resources(sample: ADMDatamart, tmp_path):
    """CDN cleanup removes the QMD-stem resources folder for named outputs."""
    hc = sample.generate.health_check(
        output_dir=tmp_path,
        name="cdn_named",
        full_embed=False,
        keep_temp_files=True,
    )

    assert hc == tmp_path / "HealthCheck_cdn_named.html"
    temp_dirs = list(tmp_path.glob("tmp_cdn_named_*"))
    assert len(temp_dirs) == 1
    assert not (temp_dirs[0] / "HealthCheck_files").exists()
    assert "HealthCheck_files/" not in (temp_dirs[0] / "HealthCheck_cdn_named.html").read_text(encoding="utf-8")


def test_GenerateHealthCheck_custom_categorization_with_unmatched_predictors(sample: ADMDatamart, tmp_path):
    sample.apply_predictor_categorization({"External Model": "Propensity"})

    hc = sample.generate.health_check(output_dir=tmp_path, name="CustomCategorization")

    _assert_report_path(hc, tmp_path, "HealthCheck_CustomCategorization")
    errors = check_report_for_errors(hc)
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


@pytest.mark.slow
def test_HealthCheck_full_embed(sample: ADMDatamart, tmp_path):
    """Test health check file sizes for full_embed options."""
    sizes = {}

    default = sample.generate.health_check(output_dir=tmp_path, name="default")
    sizes["default"] = default.stat().st_size

    print("Generating full_embed...")
    full_embed = sample.generate.health_check(
        output_dir=tmp_path,
        name="full_embed",
        full_embed=True,
    )
    sizes["full_embed"] = full_embed.stat().st_size

    print("Generating cdn...")
    cdn = sample.generate.health_check(
        output_dir=tmp_path,
        name="cdn",
        full_embed=False,
    )
    sizes["cdn"] = cdn.stat().st_size

    size_diff = abs(sizes["default"] - sizes["cdn"]) / sizes["cdn"]
    assert size_diff <= 0.10, f"Default is cdn, file sizes could be slightly different, got {size_diff:.1%} difference"
    _assert_plotly_resource_mode(cdn, full_embed, report_label="HealthCheck")
    _assert_full_embed_size_ratio(sizes, report_label="HealthCheck", min_ratio=3.0)


def test_ExportTables(sample: ADMDatamart, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    excel, warning_messages = sample.generate.excel_report(predictor_binning=True)
    assert excel == pathlib.Path("./Tables.xlsx")
    assert excel.exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
        "predictors_detail",
        "predictors_overview",
        "predictor_binning",
    ]


def test_ExportTables_NoBinning(sample: ADMDatamart, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    excel, warining_messages = sample.generate.excel_report(predictor_binning=False)
    assert excel == pathlib.Path("./Tables.xlsx")
    assert pathlib.Path(excel).exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
        "predictors_detail",
        "predictors_overview",
    ]


def test_GenerateHealthCheck_ModelDataOnly(
    sample_without_predictor_binning: ADMDatamart,
    tmp_path,
):
    hc = sample_without_predictor_binning.generate.health_check(output_dir=tmp_path, name="MyOrg")
    _assert_report_path(hc, tmp_path, "HealthCheck_MyOrg")
    errors = check_report_for_errors(hc)
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_GenerateHealthCheck_PredictionData(
    sample: ADMDatamart,
    sample_prediction_data: Prediction,
    tmp_path,
):
    hc = sample.generate.health_check(
        output_dir=tmp_path,
        prediction=sample_prediction_data,
        name="WithPredictions",
    )
    _assert_report_path(hc, tmp_path, "HealthCheck_WithPredictions")
    errors = check_report_for_errors(hc)
    assert len(errors) == 0, "HealthCheck report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


def test_ExportTables_ModelDataOnly(sample_without_predictor_binning: ADMDatamart, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    excel, warning_messages = sample_without_predictor_binning.generate.excel_report(
        name="ModelTables.xlsx",
        predictor_binning=True,
    )
    assert excel == pathlib.Path("ModelTables.xlsx")
    assert pathlib.Path(excel).exists()
    spreadsheet = load_workbook(excel)
    assert spreadsheet.sheetnames == [
        "adm_models",
    ]


def test_GenerateModelReport(sample: ADMDatamart, tmp_path):
    report = sample.generate.model_reports(
        model_ids=["bd70a915-697a-5d43-ab2c-53b0557c85a0"],
        output_dir=tmp_path,
        name="MyOrg",
        only_active_predictors=True,
    )
    _assert_report_path(report, tmp_path, "ModelReport_MyOrg_bd70a915-697a-5d43-ab2c-53b0557c85a0")
    errors = check_report_for_errors(report)
    assert len(errors) == 0, "Model report contains errors:\n" + "\n".join(f"  - {e}" for e in errors)


@pytest.mark.slow
def test_ModelReport_full_embed(sample: ADMDatamart, tmp_path):
    """Test model report file sizes for full_embed options."""
    model_id = "bd70a915-697a-5d43-ab2c-53b0557c85a0"
    sizes = {}

    default = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="default",
    )
    sizes["default"] = default.stat().st_size

    print("Generating full_embed...")
    full_embed = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="full_embed",
        full_embed=True,
    )
    sizes["full_embed"] = full_embed.stat().st_size

    cdn = sample.generate.model_reports(
        model_ids=[model_id],
        output_dir=tmp_path,
        name="cdn",
        full_embed=False,
    )
    sizes["cdn"] = cdn.stat().st_size

    size_diff = abs(sizes["default"] - sizes["cdn"]) / sizes["cdn"]
    assert size_diff <= 0.10, f"Default is cdn and sizes should be very close, got {size_diff:.1%} difference"
    _assert_plotly_resource_mode(cdn, full_embed, report_label="ModelReport")
    _assert_full_embed_size_ratio(sizes, report_label="ModelReport", min_ratio=3.0)


def test_GenerateHealthCheck_CustomQmdFile(sample: ADMDatamart, tmp_path):
    """Test health_check with custom qmd_file argument"""
    custom_qmd = tmp_path / "custom_health.qmd"
    custom_qmd.write_text("""
---
title: "Custom Health Check Report"
format: html
params:
  title: "Custom Title"
---

# Custom Health Check

This is a custom health check template.

Test parameters: {{< meta params.title >}}
""")

    hc = sample.generate.health_check(output_dir=tmp_path, name="CustomTemplate", qmd_file=custom_qmd)
    _assert_report_path(hc, tmp_path, "HealthCheck_CustomTemplate")


def test_GenerateModelReport_CustomQmdFile(sample: ADMDatamart, tmp_path):
    """Test model_reports with custom qmd_file argument"""
    custom_qmd = tmp_path / "custom_model.qmd"
    custom_qmd.write_text("""
---
title: "Custom Model Report"
format: html
params:
  title: "Custom Model Title"
  model_id: "test"
---

# Custom Model Report

This is a custom model report template.

Model ID: {{< meta params.model_id >}}
""")

    report = sample.generate.model_reports(
        model_ids=["bd70a915-697a-5d43-ab2c-53b0557c85a0"],
        output_dir=tmp_path,
        name="CustomModel",
        qmd_file=custom_qmd,
    )
    _assert_report_path(
        report,
        tmp_path,
        "ModelReport_CustomModel_bd70a915-697a-5d43-ab2c-53b0557c85a0",
    )
